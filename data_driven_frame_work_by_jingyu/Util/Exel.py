#encoding=utf-8
from openpyxl import *
from openpyxl.styles import Font

from Util.FormatTime import date_time_chinese


#openpyxl用的是2.4.5版本
class ParseExcel(object):
    def __init__(self,excel_file_path):
        self.excel_file_path=excel_file_path
        self.workbook=load_workbook(excel_file_path)
        self.font=Font(color=None)
        self.colorDict={"red":'FFFF3030',"green":'FF008B00'}
        self.sheet=self.workbook.active


    #设置当前要操作的sheet对象，使用index来获取相应的sheet,索引从0开始
    def set_sheet_by_index(self,sheet_index):
        self.sheet = self.get_sheet_by_index(sheet_index)

    # 设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet
    def set_sheet_by_name(self,sheet_name):
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)

    #获取当前默认sheet的名字
    def get_default_name(self):
        return self.sheet.title

    #通过sheet名称获取sheet对象
    def get_sheet_by_name(self,sheet_name):
        self.sheet=self.workbook.get_sheet_by_name(sheet_name)
        return self.sheet

    # 通过sheet 索引获取sheet对象
    def get_sheet_by_index(self,sheet_index):
        sheet_name=self.workbook.get_sheet_names()[sheet_index]
        self.sheet=self.get_sheet_by_name(sheet_name)
        return self.sheet

    #获取默认sheet中最大的行数
    def get_max_row_no(self):
        return self.sheet.max_row

    #获取默认 sheet 的最大列数
    def get_max_col_no(self):
        return self.sheet.max_column

    #获取默认sheet的最小（起始）行号
    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet的最小（起始）列号
    def get_min_col_no(self):
        return self.sheet.min_column


    # 获取默认 sheet 的所有行对象，
    def get_all_rows(self):
        #print dir(self.sheet)
        rows=[]
        for row in self.sheet.iter_rows():
            rows.append(row)
        #return list(self.sheet.rows()) #这样的返回也可以
        return rows

    #获取默认sheet中的所有列对象
    def get_all_cols(self):
        cols = []
        for col in self.sheet.iter_cols():
            cols.append(col)
        # return list(self.sheet.columns()) #这样的返回也可以
        return cols

    #从默认sheet中获取某一行，第一行从0开始
    def get_single_row(self,row_no):
        return self.get_all_rows()[row_no]

    # #从默认sheet中获取某一列，第一列从0开始
    def get_single_col(self,col_no):
        return self.get_all_cols()[col_no]

    #从默认sheet中，通过行号和列号获取指定的单元格，注意行号和列号从1开始
    def get_cell(self,row_no,col_no):
        return self.sheet.cell(row=row_no,column=col_no)

    # 从默认sheet中，通过行号和列号获取指定的单元格中的内容，注意行号和列号从1开始
    def get_cell_content(self,row_no,col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入指定内容，注意行号和列号从1开始
    # 调用此方法的时候，excel不要处于打开状态
    def write_cell_content(self,row_no,col_no,content,font=None):
        self.sheet.cell(row=row_no, column=col_no).value=content
        self.workbook.save(self.excel_file_path)

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期，注意行号和列号从1开始
    #调用此方法的时候，excel不要处于打开状态
    def write_cell_current_time(self,row_no,col_no):
        self.sheet.cell(row=row_no, column=col_no).value =date_time_chinese()
        self.workbook.save(self.excel_file_path)

    def save_excel_file(self):
        self.workbook.save(self.excel_file_path)

if __name__=='__main__':
    pe=ParseExcel("D:\\data_driven_all\\DataDrivenFrameWork_ke\\TestData\\data.xlsx")
    #pe.set_sheet_by_index(0)
    #print pe.get_default_name()
   # print pe.get_default_name()
    print pe.get_sheet_by_index(0)
    print 'max row:',pe.get_max_row_no()
    '''
    print pe.get_all_rows() #获取所有行对象
    print pe.get_all_rows()[0] #获取第一行，获取某个行对象
    print pe.get_all_rows()[0][1] #获取第一行，第二列的单元格对象
    print pe.get_all_rows()[0][1].value #获取第一行，第二列单元格对象的值

    print pe.get_all_cols()  # 获取所有行对象
    print pe.get_all_cols()[0]  # 获取第一列，获取某个列对象
    print pe.get_all_cols()[0][1]  # 获取第一列的第二个单元格某个单元格对象
    print pe.get_all_cols()[0][1].value  # 获取第一列的第二个单元格的值

    print pe.get_max_col_no()
    print pe.get_min_row_no()
    print pe.get_min_col_no()
    print pe.get_all_rows()[0]
    print len(pe.get_all_rows())
    for cell in pe.get_all_rows()[0]:
        print cell.value
    print pe.get_all_cols()
    for col in pe.get_all_cols()[0]:
        print col.value
    for cell in pe.get_single_row(0):
        print cell.value
    for col in pe.get_single_col(0):
        print col.value
    print pe.get_cell(1,1)

    print pe.get_cell_content(1,1)
    pe.write_cell_content(8,8,"光荣之路")
    print pe.get_cell_content(8, 8)
    pe.write_cell_current_time(8, 8)
    print pe.get_cell_content(8, 8)
    '''